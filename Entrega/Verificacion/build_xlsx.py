import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Solver_Transporte"

NAVY = "1F3864"
LIGHT = "DCE6F1"
navy_fill = PatternFill("solid", fgColor=NAVY)
light_fill = PatternFill("solid", fgColor=LIGHT)
white_bold = Font(color="FFFFFF", bold=True)
bold = Font(bold=True)
thin = Side(style="thin", color="999999")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def set_cell(cell, value, font=None, fill=None, align="center", fmt=None, brd=True):
    c = ws[cell]
    c.value = value
    if font: c.font = font
    if fill: c.fill = fill
    c.alignment = Alignment(horizontal=align, vertical="center")
    if fmt: c.number_format = fmt
    if brd: c.border = border
    return c

ws["A1"] = "MODELO DE TRANSPORTE — HOJA DE SOLVER (Excel Solver / LibreOffice Calc Solver)"
ws["A1"].font = Font(bold=True, size=13, color=NAVY)
ws.merge_cells("A1:J1")

ws["A2"] = "Variables de decisión Xij: unidades enviadas del centro i (fila) al punto de venta j (columna). Estas son las CELDAS CAMBIANTES de Solver."
ws["A2"].font = Font(italic=True, size=9, color="555555")
ws.merge_cells("A2:J2")

# ---- Decision variable block (rows 4-7) ----
set_cell("A3", "Envíos (Xij)", white_bold, navy_fill)
puntos = ["Punto 1", "Punto 2", "Punto 3", "Punto 4", "Punto 5"]
for i, pname in enumerate(puntos):
    set_cell(get_column_letter(2 + i) + "3", pname, white_bold, navy_fill)
set_cell("G3", "Disponibilidad", white_bold, navy_fill)
set_cell("H3", "Usado (=SUMA fila)", white_bold, navy_fill)
set_cell("I3", "Restricción", white_bold, navy_fill)

centros = ["Centro A", "Centro B", "Centro C"]
disponibilidad = [100, 150, 200]
for r, (cname, disp) in enumerate(zip(centros, disponibilidad)):
    row = 4 + r
    set_cell(f"A{row}", cname, bold, light_fill)
    for c in range(5):
        col = get_column_letter(2 + c)
        set_cell(f"{col}{row}", 0, fmt="0")
    set_cell(f"G{row}", disp, fmt="0")
    set_cell(f"H{row}", f"=SUM(B{row}:F{row})", fmt="0")
    set_cell(f"I{row}", f"H{row}<=G{row}", align="left", brd=False)

set_cell("A7", "Demanda", bold, light_fill)
demanda = [80, 90, 120, 60, 100]
for c, d in enumerate(demanda):
    col = get_column_letter(2 + c)
    set_cell(f"{col}7", d, fmt="0")

set_cell("A8", "Recibido (=SUMA col.)", bold, light_fill)
for c in range(5):
    col = get_column_letter(2 + c)
    set_cell(f"{col}8", f"=SUM({col}4:{col}6)", fmt="0")
set_cell("H8", "Restricción: fila 8 = fila 7 (una por columna)", align="left", brd=False)

# ---- Cost matrix block (rows 11-14) ----
set_cell("A10", "Costos unitarios de transporte (cij)", white_bold, navy_fill)
ws.merge_cells("A10:F10")
for i, pname in enumerate(puntos):
    set_cell(get_column_letter(2 + i) + "11", pname, white_bold, navy_fill)
set_cell("A11", "", white_bold, navy_fill)
costos = {
    "Centro A": [4, 6, 8, 10, 12],
    "Centro B": [5, 7, 9, 11, 13],
    "Centro C": [6, 8, 10, 12, 14],
}
for r, cname in enumerate(centros):
    row = 12 + r
    set_cell(f"A{row}", cname, bold, light_fill)
    for c, val in enumerate(costos[cname]):
        col = get_column_letter(2 + c)
        set_cell(f"{col}{row}", val, fmt="0")

# ---- Objective cell ----
set_cell("A17", "FUNCIÓN OBJETIVO (celda objetivo de Solver)", white_bold, navy_fill)
ws.merge_cells("A17:C17")
set_cell("D17", "Minimizar Z =", bold, align="right", brd=False)
set_cell("E17", "=SUMPRODUCT(B4:F6,B12:F16)".replace("F16", "F14"), Font(bold=True, size=12, color="C00000"), fmt="#,##0")
ws.merge_cells("E17:F17")

ws.column_dimensions["A"].width = 24
for col in "BCDEFGHI":
    ws.column_dimensions[col].width = 14

from openpyxl.worksheet.properties import PageSetupProperties
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
ws.print_area = "A1:J20"

wb.save("Modelo_Transporte_Solver.xlsx")
print("xlsx creado")
